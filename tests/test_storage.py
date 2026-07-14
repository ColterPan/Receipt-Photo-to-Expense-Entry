from datetime import datetime

from openpyxl import Workbook, load_workbook

from receipt_expense.models import ExpenseRow
from receipt_expense.storage import HEADER, SHEET_NAME, append_expense_row, archive_image


def sample_row(**overrides) -> ExpenseRow:
    data = dict(
        date="2026-07-10",
        vendor="Coffee House",
        amount=12.50,
        currency="USD",
        category="Meals",
        status="OK",
        notes="",
        image_path="data/archive/2026/07/x.jpg",
        processed_at=datetime(2026, 7, 14, 12, 0, 0),
    )
    data.update(overrides)
    return ExpenseRow(**data)


def test_creates_workbook_with_header(tmp_path):
    f = tmp_path / "expenses.xlsx"
    n = append_expense_row(sample_row(), f)
    assert n == 2

    wb = load_workbook(f)
    sheet = wb[SHEET_NAME]
    assert [c.value for c in sheet[1]] == HEADER
    assert sheet.cell(row=2, column=2).value == "Coffee House"
    assert sheet.cell(row=2, column=3).value == 12.50


def test_appends_below_existing_rows(tmp_path):
    f = tmp_path / "expenses.xlsx"
    append_expense_row(sample_row(), f)
    n = append_expense_row(sample_row(vendor="Taxi Co", category="Transport"), f)
    assert n == 3

    sheet = load_workbook(f)[SHEET_NAME]
    assert sheet.cell(row=3, column=2).value == "Taxi Co"


def test_appends_to_existing_workbook_with_other_sheets(tmp_path):
    f = tmp_path / "expenses.xlsx"
    wb = Workbook()
    wb.active.title = "Budget"
    wb.active.append(["existing", "data"])
    wb.save(f)

    append_expense_row(sample_row(), f)

    wb2 = load_workbook(f)
    assert "Budget" in wb2.sheetnames  # untouched
    assert [c.value for c in wb2[SHEET_NAME][1]] == HEADER


def test_none_amount_written_as_blank_cell(tmp_path):
    f = tmp_path / "expenses.xlsx"
    append_expense_row(sample_row(amount=None, status="NEEDS REVIEW"), f)
    sheet = load_workbook(f)[SHEET_NAME]
    assert sheet.cell(row=2, column=3).value is None
    assert sheet.cell(row=2, column=6).value == "NEEDS REVIEW"


def test_archive_image_paths_and_content(tmp_path):
    path = archive_image(b"fake-jpeg", tmp_path, "2026-07-10", "Coffee House & Co!")
    assert path.read_bytes() == b"fake-jpeg"
    assert path.parent == tmp_path / "2026" / "07"
    assert path.name.startswith("2026-07-10-coffee-house-co-")
    assert path.suffix == ".jpg"


def test_archive_image_bad_date_falls_back_to_today(tmp_path):
    path = archive_image(b"x", tmp_path, "", "Vendor")
    assert path.exists()
