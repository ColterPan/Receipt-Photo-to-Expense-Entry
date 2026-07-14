import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from PIL import Image

from receipt_expense import extractor
from receipt_expense.app import app
from receipt_expense.config import get_settings
from receipt_expense.models import ReceiptExtraction


@pytest.fixture()
def client(tmp_path, monkeypatch):
    monkeypatch.setenv("EXPENSE_FILE", str(tmp_path / "expenses.xlsx"))
    monkeypatch.setenv("ARCHIVE_DIR", str(tmp_path / "archive"))
    monkeypatch.setenv("ANTHROPIC_API_KEY", "test-key")
    get_settings.cache_clear()
    yield TestClient(app)
    get_settings.cache_clear()


def receipt_photo() -> bytes:
    buf = io.BytesIO()
    Image.new("RGB", (600, 800), (255, 255, 255)).save(buf, format="JPEG")
    return buf.getvalue()


def test_process_receipt_happy_path(client, tmp_path, monkeypatch):
    monkeypatch.setattr(
        extractor,
        "extract_receipt",
        lambda jpeg, categories, client=None: ReceiptExtraction(
            vendor="Coffee House", amount=12.5, currency="USD",
            date="2026-07-10", category="Meals", confidence="high", notes="",
        ),
    )

    r = client.post("/api/receipts", files={"file": ("r.jpg", receipt_photo(), "image/jpeg")})
    assert r.status_code == 200
    body = r.json()
    assert body["status"] == "OK"
    assert body["vendor"] == "Coffee House"

    sheet = load_workbook(tmp_path / "expenses.xlsx")["Expenses"]
    assert sheet.max_row == 2
    archived = list((tmp_path / "archive").rglob("*.jpg"))
    assert len(archived) == 1


def test_uncertain_extraction_is_flagged(client, tmp_path, monkeypatch):
    monkeypatch.setattr(
        extractor,
        "extract_receipt",
        lambda jpeg, categories, client=None: ReceiptExtraction(
            vendor="Blurry Mart", amount=None, currency=None,
            date=None, category="Uncategorized", confidence="low",
            notes="total is unreadable",
        ),
    )

    r = client.post("/api/receipts", files={"file": ("r.jpg", receipt_photo(), "image/jpeg")})
    assert r.status_code == 200
    body = r.json()
    assert body["status"] == "NEEDS REVIEW"
    assert body["amount"] is None
    assert "unreadable" in body["notes"]


def test_non_image_upload_is_rejected(client):
    r = client.post("/api/receipts", files={"file": ("r.txt", b"not an image", "text/plain")})
    assert r.status_code == 400


def test_extraction_error_maps_to_gateway_error(client, monkeypatch):
    def boom(jpeg, categories, client=None):
        raise extractor.ExtractionError("API key missing")

    monkeypatch.setattr(extractor, "extract_receipt", boom)
    r = client.post("/api/receipts", files={"file": ("r.jpg", receipt_photo(), "image/jpeg")})
    assert r.status_code == 502
    assert "API key" in r.json()["detail"]


def test_info_endpoint(client):
    r = client.get("/api/info")
    assert r.status_code == 200
    body = r.json()
    assert body["api_key_configured"] is True
    assert "Meals" in body["categories"]
