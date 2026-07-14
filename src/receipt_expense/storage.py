"""Persist expense rows to the Excel workbook and archive receipt photos."""

import logging
import re
import threading
import uuid
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import ExpenseRow

logger = logging.getLogger(__name__)

SHEET_NAME = "Expenses"
HEADER = [
    "Date", "Vendor", "Amount", "Currency", "Category",
    "Status", "Notes", "Receipt Image", "Processed At",
]

# Single-user local app: guard against two browser tabs submitting at once.
# Concurrent writes from multiple *processes* are not supported (documented in README).
_write_lock = threading.Lock()


class ExpenseFileLockedError(RuntimeError):
    """The workbook could not be saved (usually because it is open in Excel)."""


def archive_image(jpeg_bytes: bytes, archive_dir: Path, row_date: str, vendor: str) -> Path:
    """Save the normalized receipt photo under archive/YYYY/MM/ and return its path."""
    when = _safe_date(row_date)
    folder = archive_dir / f"{when:%Y}" / f"{when:%m}"
    folder.mkdir(parents=True, exist_ok=True)

    slug = re.sub(r"[^a-z0-9]+", "-", vendor.lower()).strip("-") or "receipt"
    path = folder / f"{when:%Y-%m-%d}-{slug[:40]}-{uuid.uuid4().hex[:8]}.jpg"
    path.write_bytes(jpeg_bytes)
    logger.info("Archived receipt image to %s", path)
    return path


def append_expense_row(row: ExpenseRow, expense_file: Path) -> int:
    """Append a row to the expense workbook, creating file/sheet + header if missing.

    Returns the 1-based row number that was written.
    """
    with _write_lock:
        expense_file.parent.mkdir(parents=True, exist_ok=True)

        if expense_file.exists():
            workbook = load_workbook(expense_file)
            if SHEET_NAME in workbook.sheetnames:
                sheet = workbook[SHEET_NAME]
            else:
                sheet = workbook.create_sheet(SHEET_NAME)
                sheet.append(HEADER)
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = SHEET_NAME
            sheet.append(HEADER)

        _ensure_header(sheet)
        sheet.append([
            row.date,
            row.vendor,
            row.amount,
            row.currency,
            row.category,
            row.status,
            row.notes,
            row.image_path,
            row.processed_at.strftime("%Y-%m-%d %H:%M:%S"),
        ])
        row_number = sheet.max_row

        try:
            workbook.save(expense_file)
        except PermissionError as exc:
            raise ExpenseFileLockedError(
                f"Could not save {expense_file}. Close it in Excel and try again."
            ) from exc

    logger.info("Appended expense row %d to %s", row_number, expense_file)
    return row_number


def _ensure_header(sheet: Worksheet) -> None:
    """An existing but empty sheet still gets the header row."""
    if sheet.max_row == 1 and all(cell.value is None for cell in sheet[1]):
        sheet.append(HEADER)
        sheet.delete_rows(1)


def _safe_date(row_date: str) -> datetime:
    try:
        return datetime.strptime(row_date, "%Y-%m-%d")
    except ValueError:
        return datetime.now()
